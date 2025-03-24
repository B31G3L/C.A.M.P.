"""
Modal-Fenster zum Importieren von Daten mit Tabs für BW Tool Daten und Projektdaten
"""
import os
import customtkinter as ctk
from typing import Dict, Callable, Optional, List
import tkinter as tk
from tkinter import filedialog, messagebox
import csv
import json
import shutil
import datetime
import pandas as pd
import re


class DragDropFrame(ctk.CTkFrame):
    """
    Ein Frame, das Drag & Drop von Dateien ermöglicht
    """
    def __init__(self, master, callback, accepted_files=None, **kwargs):
        """
        Initialisiert den Drag & Drop Frame
        
        Args:
            master: Das übergeordnete Widget
            callback: Funktion, die aufgerufen wird, wenn eine Datei fallen gelassen wird
            accepted_files: Liste der akzeptierten Dateierweiterungen (z.B. ['.xls', '.xlsx'])
        """
        super().__init__(master, **kwargs)
        
        self.callback = callback
        self.accepted_files = accepted_files or ['.xls', '.xlsx']
        
        # Aussehen des Frames
        self.configure(
            fg_color=("gray90", "gray20"),
            corner_radius=10,
            border_width=2,
            border_color=("gray70", "gray40")
        )
        
        # Text und Icon für den Drop-Bereich
        self.text_label = ctk.CTkLabel(
            self,
            text="Datei hier hineinziehen oder klicken zum Auswählen",
            font=ctk.CTkFont(size=14)
        )
        self.text_label.pack(expand=True, pady=(20, 10))
        
        self.info_label = ctk.CTkLabel(
            self,
            text=f"Akzeptierte Dateitypen: {', '.join(self.accepted_files)}",
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray70")
        )
        self.info_label.pack(expand=True, pady=(0, 20))
        
        # Erhöhe den Frame so, dass er gut als Drop-Ziel funktioniert
        self.configure(height=150)
        
        # Ereignisbehandlung für Mausklicks
        self.bind("<Button-1>", self._on_click)
        self.text_label.bind("<Button-1>", self._on_click)
        self.info_label.bind("<Button-1>", self._on_click)
        
        # Richte Drag & Drop ein für Windows (TkDND)
        self._setup_tkdnd()
    
    def _setup_tkdnd(self):
        """Richtet TkDND für Drag & Drop-Funktionalität ein, falls verfügbar"""
        try:
            self.winfo_toplevel().tk.call('package', 'require', 'tkdnd')
            self.winfo_toplevel().tk.call('tkdnd::drop_target', 'register', self._get_widget_path(), self.accepted_files)
            self.bind('<<Drop>>', self._on_drop)
            
            # Visuelle Rückmeldung beim Überfahren
            self.bind('<<DropEnter>>', lambda e: self.configure(border_color=("blue", "light blue")))
            self.bind('<<DropLeave>>', lambda e: self.configure(border_color=("gray70", "gray40")))
        except tk.TclError:
            # TkDND ist nicht verfügbar, füge eine Info hinzu
            self.info_label.configure(text="Drag & Drop nicht verfügbar, bitte klicken zum Auswählen")
    
    def _get_widget_path(self):
        """Gibt den TK-Widget-Pfad für TkDND zurück"""
        return self.winfo_toplevel().register(self)
    
    def _on_click(self, event):
        """Wird aufgerufen, wenn auf den Frame geklickt wird"""
        filetypes = [("Excel-Dateien", "*.xls *.xlsx"), ("Alle Dateien", "*.*")]
        filename = filedialog.askopenfilename(
            title="Excel-Datei auswählen",
            filetypes=filetypes
        )
        
        if filename:
            self.callback(filename)
    
    def _on_drop(self, event):
        """Wird aufgerufen, wenn eine Datei auf den Frame gezogen wird"""
        self.configure(border_color=("gray70", "gray40"))  # Zurücksetzen der Rahmenfarbe
        
        # Prüfe, ob die Datei einen der akzeptierten Dateitypen hat
        filename = event.data
        _, file_extension = os.path.splitext(filename)
        
        if file_extension.lower() in self.accepted_files:
            self.callback(filename)
        else:
            messagebox.showerror(
                "Ungültiger Dateityp",
                f"Die Datei {os.path.basename(filename)} hat einen ungültigen Dateityp. "
                f"Akzeptierte Typen: {', '.join(self.accepted_files)}",
                parent=self.winfo_toplevel()
            )


class ProjectMemberSelectionDialog(ctk.CTkToplevel):
    """Dialog zur Auswahl eines Projektmitglieds"""
    
    def __init__(self, master, projects_data, on_select=None):
        """
        Initialisiert den Dialog
        
        Args:
            master: Das übergeordnete Widget
            projects_data: Die Projektdaten aus der project.json
            on_select: Callback-Funktion, die aufgerufen wird, wenn ein Mitglied ausgewählt wird
        """
        super().__init__(master)
        
        # Fenster zunächst ausblenden
        self.withdraw()
        
        # Fenster-Konfiguration
        self.title("Projektmitglied auswählen")
        self.geometry("500x400")
        self.resizable(False, False)
        self.transient(master)  # Macht das Fenster zum Modal-Fenster
        self.grab_set()         # Blockiert Interaktion mit dem Hauptfenster
        
        # Attribute
        self.projects_data = projects_data
        self.on_select = on_select
        self.selected_member = None
        
        # UI erstellen
        self._create_widgets()
        self._setup_layout()
        self._load_projects()
        
        # Fenster zentrieren und dann anzeigen
        self.center_window()
        
        # Nach einer kurzen Verzögerung anzeigen
        self.after(100, self.deiconify)
    
    def _create_widgets(self):
        """Erstellt alle UI-Elemente für den Dialog"""
        # Titel-Label
        self.title_label = ctk.CTkLabel(
            self, 
            text="Wählen Sie ein Projektmitglied aus",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        
        # Projekt-Auswahl
        self.project_frame = ctk.CTkFrame(self)
        self.project_label = ctk.CTkLabel(
            self.project_frame,
            text="Projekt:",
            width=100
        )
        self.project_var = ctk.StringVar(value="")
        self.project_dropdown = ctk.CTkOptionMenu(
            self.project_frame,
            values=[""],
            variable=self.project_var,
            command=self._on_project_selected,
            width=300
        )
        
        # Mitglieder-Listen-Frame
        self.members_label = ctk.CTkLabel(
            self,
            text="Projektmitglieder:",
            anchor="w"
        )
        
        # Scrollbare Liste für Mitglieder
        self.members_list_frame = ctk.CTkScrollableFrame(
            self,
            width=450,
            height=250
        )
        
        # Info-Label
        self.info_label = ctk.CTkLabel(
            self,
            text="",
            text_color=("gray60", "gray70")
        )
        
        # Button-Frame
        self.button_frame = ctk.CTkFrame(self)
        
        # Abbrechen-Button
        self.cancel_button = ctk.CTkButton(
            self.button_frame,
            text="Abbrechen",
            command=self.destroy,
            width=120,
            fg_color=("gray70", "gray30"),
            hover_color=("gray60", "gray40")
        )
        
        # Auswählen-Button (initial deaktiviert)
        self.select_button = ctk.CTkButton(
            self.button_frame,
            text="Auswählen",
            command=self._on_select_clicked,
            width=120,
            state="disabled"
        )
    
    def _setup_layout(self):
        """Richtet das Layout der UI-Elemente ein"""
        # Titel-Label
        self.title_label.pack(pady=(20, 15))
        
        # Projekt-Auswahl
        self.project_frame.pack(fill="x", padx=20, pady=(0, 15))
        self.project_label.pack(side="left")
        self.project_dropdown.pack(side="left", padx=(10, 0))
        
        # Mitglieder-Label
        self.members_label.pack(anchor="w", padx=20, pady=(0, 5))
        
        # Mitglieder-Liste
        self.members_list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Info-Label
        self.info_label.pack(padx=20, pady=(0, 10))
        
        # Button-Frame
        self.button_frame.pack(fill="x", padx=20, pady=(0, 20))
        self.cancel_button.pack(side="left", padx=(0, 10))
        self.select_button.pack(side="right")
    
    def _load_projects(self):
        """Lädt die Projekte in das Dropdown"""
        if not self.projects_data or "projekte" not in self.projects_data:
            self.info_label.configure(
                text="Keine Projekte gefunden. Bitte importieren Sie zuerst Projektdaten.",
                text_color="red"
            )
            return
        
        # Projekte extrahieren
        projects = self.projects_data.get("projekte", [])
        project_names = [project.get("name", "") for project in projects if project.get("name")]
        
        if not project_names:
            self.info_label.configure(
                text="Keine gültigen Projekte gefunden.",
                text_color="red"
            )
            return
        
        # Dropdown aktualisieren
        self.project_var.set(project_names[0])
        self.project_dropdown.configure(values=project_names)
        
        # Zeige Mitglieder des ersten Projekts
        self._on_project_selected(project_names[0])
    
    def _on_project_selected(self, project_name):
        """Wird aufgerufen, wenn ein Projekt ausgewählt wird"""
        # Bisherige Auswahl zurücksetzen
        self.selected_member = None
        self.select_button.configure(state="disabled")
        
        # Bisherige Mitgliederliste löschen
        for widget in self.members_list_frame.winfo_children():
            widget.destroy()
        
        # Finde das ausgewählte Projekt
        selected_project = None
        for project in self.projects_data.get("projekte", []):
            if project.get("name") == project_name:
                selected_project = project
                break
        
        if not selected_project:
            self.info_label.configure(
                text=f"Projekt '{project_name}' nicht gefunden.",
                text_color="red"
            )
            return
        
        # Mitglieder des Projekts extrahieren
        members = selected_project.get("teilnehmer", [])
        
        if not members:
            self.info_label.configure(
                text=f"Keine Mitglieder im Projekt '{project_name}' gefunden.",
                text_color="red"
            )
            return
        
        # Info-Label aktualisieren
        self.info_label.configure(
            text=f"{len(members)} Mitglieder gefunden. Bitte wählen Sie ein Mitglied aus.",
            text_color=("gray60", "gray70")
        )
        
        # Mitglieder anzeigen
        for i, member in enumerate(members):
            member_id = member.get("id", "")
            member_name = member.get("name", "")
            member_role = member.get("rolle", "")
            
            # Erstelle ein Frame für den Mitgliedseintrag
            member_frame = ctk.CTkFrame(self.members_list_frame)
            member_frame.pack(fill="x", padx=5, pady=2)
            
            # Hintergrundfarbe abwechseln
            if i % 2 == 0:
                member_frame.configure(fg_color=("gray90", "gray20"))
            
            # Speichere die Mitgliedsdaten im Frame
            member_frame.member_data = member
            
            # Erstelle Labels für Name und Rolle
            name_label = ctk.CTkLabel(
                member_frame,
                text=member_name,
                anchor="w",
                width=200
            )
            
            role_label = ctk.CTkLabel(
                member_frame,
                text=member_role,
                anchor="w",
                width=150
            )
            
            id_label = ctk.CTkLabel(
                member_frame,
                text=member_id,
                anchor="w",
                width=100
            )
            
            
            # Platziere die Labels
            name_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
            role_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")
            id_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")
            
            member_frame.bind("<Button-1>", lambda e, f=member_frame: self._on_member_clicked(f))
            # Doppelklick zur direkten Auswahl
            member_frame.bind("<Double-Button-1>", lambda e, f=member_frame: self._on_member_double_click(f))
            name_label.bind("<Button-1>", lambda e, f=member_frame: self._on_member_clicked(f))
            name_label.bind("<Double-Button-1>", lambda e, f=member_frame: self._on_member_double_click(f))
            role_label.bind("<Button-1>", lambda e, f=member_frame: self._on_member_clicked(f))
            role_label.bind("<Double-Button-1>", lambda e, f=member_frame: self._on_member_double_click(f))
            id_label.bind("<Button-1>", lambda e, f=member_frame: self._on_member_clicked(f))
            id_label.bind("<Double-Button-1>", lambda e, f=member_frame: self._on_member_double_click(f))

    def _on_member_double_click(self, member_frame):
                """Wird aufgerufen, wenn auf ein Mitglied doppelgeklickt wird"""
                # Wähle das Mitglied aus
                self._on_member_clicked(member_frame)
                
                # Bestätige die Auswahl sofort, als würde der Auswählen-Button geklickt
                if self.selected_member:
                    self._on_select_clicked()     
    
    def _on_member_clicked(self, member_frame):
        """Wird aufgerufen, wenn ein Mitglied angeklickt wird"""
        # Bisherige Auswahl zurücksetzen
        for widget in self.members_list_frame.winfo_children():
            if hasattr(widget, 'member_data'):
                i = list(self.members_list_frame.winfo_children()).index(widget)
                if i % 2 == 0:
                    widget.configure(fg_color=("gray90", "gray20"))
                else:
                    widget.configure(fg_color="transparent")
        
        # Diese Zeile als ausgewählt markieren
        member_frame.configure(fg_color=("light blue", "dark blue"))
        
        # Mitgliedsdaten speichern
        self.selected_member = member_frame.member_data
        
        # Auswählen-Button aktivieren
        self.select_button.configure(state="normal")
        
        # Info-Label aktualisieren
        member_name = self.selected_member.get("name", "")
        self.info_label.configure(
            text=f"Ausgewählt: {member_name}",
            text_color=("gray60", "gray70")
        )
    
    def _on_select_clicked(self):
        """Wird aufgerufen, wenn der Auswählen-Button geklickt wird"""
        if self.selected_member and self.on_select:
            self.on_select(self.selected_member)
        
        # Dialog schließen
        self.destroy()
    
    def center_window(self):
        """Zentriert das Fenster auf dem Bildschirm"""
        self.update_idletasks()
        
        # Fenstergröße
        width = self.winfo_width()
        height = self.winfo_height()
        
        # Bildschirmgröße
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Position berechnen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # Fenster positionieren
        self.geometry(f"{width}x{height}+{x}+{y}")


class ImportModal(ctk.CTkToplevel):
    """
    Modal-Fenster zum Importieren von Daten mit Tabs für BW Tool Daten und Projektdaten
    """
    def __init__(self, master, callbacks: Optional[Dict[str, Callable]] = None):
        """
        Initialisiert das Modal-Fenster
        
        Args:
            master: Das übergeordnete Widget
            callbacks: Dictionary mit Callback-Funktionen
        """
        super().__init__(master)
        
        # Fenster zunächst ausblenden
        self.withdraw()
        
        # Fenster-Konfiguration
        self.title("Daten importieren")
        self.geometry("700x550")
        self.minsize(600, 450)  # Minimale Fenstergröße festlegen
        self.transient(master)  # Macht das Fenster zum Modal-Fenster
        self.grab_set()         # Blockiert Interaktion mit dem Hauptfenster
        
        # Attribute
        self.callbacks = callbacks or {}
        self.selected_file = None
        self.import_status = None
        
        # Neue Attribute für die Arbeitszeiterfassung
        self.ar_selected_file = None
        self.selected_member = None
        
        # UI erstellen
        self._create_widgets()
        self._setup_layout()
        
        # Fenster zentrieren und dann anzeigen
        self.center_window()
        
        # Nach einer kurzen Verzögerung anzeigen (warten bis UI gerendert ist)
        self.after(100, self.deiconify)
        
        # Stelle sicher, dass das Fenster tatsächlich angezeigt wird
        self.attributes('-topmost', True)
        self.after(200, lambda: self.attributes('-topmost', False))
    
    def _create_widgets(self):
        """Erstellt alle UI-Elemente für das Modal"""
        # Titel-Frame
        self.title_frame = ctk.CTkFrame(self)
        self.title_label = ctk.CTkLabel(
            self.title_frame, 
            text="Daten importieren",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        
        # Status-Label
        self.status_frame = ctk.CTkFrame(self)
        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="",
            font=ctk.CTkFont(size=12),
            text_color=("gray60", "gray70")
        )
        
        # Tab-View für die verschiedenen Import-Optionen
        self.tab_view = ctk.CTkTabview(self)
        
        # Tabs erstellen
        self.bw_tool_tab = self.tab_view.add("BW Tool Daten")
        self.ar_calendar_tab = self.tab_view.add("Arbeitszeiterfassung")
        self.projekt_tab = self.tab_view.add("Projektdaten")
        
        # Inhalte für BW Tool Daten-Tab
        self._create_bw_tool_tab()
        
        # Inhalte für Arbeitszeiterfassung-Tab
        self._create_timesheet_tab()
        
        # Inhalte für Projektdaten-Tab
        self._create_projekt_tab()
        
        # Button-Frame (unten)
        self.button_frame = ctk.CTkFrame(self)
        self.close_button = ctk.CTkButton(
            self.button_frame, 
            text="Schließen",
            command=self.destroy,
            width=120
        )
    
    def _create_bw_tool_tab(self):
        """Erstellt die Inhalte für den BW Tool Daten-Tab"""
        # Container-Frame für den Tab
        container = ctk.CTkFrame(self.bw_tool_tab)
        
        # Info-Text
        info_text = (
            "Importieren Sie BW Tool Daten aus einer Excel-Datei (.xls, .xlsx).\n\n"
            "Die Daten werden aus der Excel-Datei extrahiert und in folgendem Format gespeichert:\n"
            "MITARBEITER_ID;DATUM;STUNDEN\n\n"
            "Es werden nur Einträge berücksichtigt, die in Spalte 11 'FCAST', 'Forecast' oder 'FORECAST' enthalten."
        )
        
        info_label = ctk.CTkLabel(
            container,
            text=info_text,
            justify="left",
            anchor="w",
            wraplength=600
        )
        
        # Drag & Drop Frame
        self.drop_frame = DragDropFrame(
            container,
            callback=self._on_file_selected,
            accepted_files=['.xls', '.xlsx']
        )
        
        # Dateiauswahl-Frame (für Anzeige des ausgewählten Dateinamens)
        file_frame = ctk.CTkFrame(container)
        
        # Label für ausgewählte Datei
        file_label = ctk.CTkLabel(
            file_frame,
            text="Ausgewählte Datei:",
            font=ctk.CTkFont(weight="bold"),
            width=120
        )
        
        # Anzeige des Dateinamens
        self.file_name_label = ctk.CTkLabel(
            file_frame,
            text="Keine Datei ausgewählt",
            anchor="w"
        )
        
        # Optionen-Frame
        options_frame = ctk.CTkFrame(container)
        
        # Optionen für den Import
        self.bw_tool_overwrite_var = ctk.BooleanVar(value=True)
        self.bw_tool_overwrite_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Bestehende Einträge überschreiben (gleiche ID und Datum)",
            variable=self.bw_tool_overwrite_var
        )
        
        self.bw_tool_backup_var = ctk.BooleanVar(value=True)
        self.bw_tool_backup_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Backup erstellen",
            variable=self.bw_tool_backup_var
        )
        
        # Import-Button
        self.bw_tool_import_button = ctk.CTkButton(
            container,
            text="BW Tool Daten importieren",
            command=self._import_bw_tool_data,
            font=ctk.CTkFont(weight="bold"),
            height=40,
            fg_color=("green3", "green4"),
            hover_color=("green4", "green3"),
            state="disabled"  # Initial deaktiviert
        )
        
        # Layout für den BW Tool Daten-Tab
        container.pack(fill="both", expand=True, padx=20, pady=20)
        info_label.pack(anchor="w", pady=(0, 15))
        
        # Drag & Drop Frame
        self.drop_frame.pack(fill="x", pady=(0, 15))
        
        # Dateiauswahl-Frame
        file_frame.pack(fill="x", pady=(0, 15))
        file_label.pack(side="left")
        self.file_name_label.pack(side="left", padx=10, fill="x", expand=True)
        
        # Optionen-Frame
        options_frame.pack(fill="x", pady=(0, 20))
        self.bw_tool_overwrite_checkbox.pack(anchor="w", pady=5)
        self.bw_tool_backup_checkbox.pack(anchor="w", pady=5)
        
        # Import-Button
        self.bw_tool_import_button.pack(fill="x", pady=(10, 0))
    
    def _create_timesheet_tab(self):
        """Erstellt die Inhalte für den Arbeitszeiterfassung-Tab"""
        # Container-Frame für den Tab
        container = ctk.CTkFrame(self.ar_calendar_tab)
        
        # Info-Text
        info_text = (
            "Importieren Sie Arbeitszeiterfassungsdaten aus einer Excel-Datei (.xlsx, .xls).\n\n"
            "Die Datei sollte Spalten für 'Date' und 'TotalHours' enthalten. "
            "Da die Datei keine Mitarbeiter-ID enthält, müssen Sie einen Mitarbeiter auswählen, "
            "dem diese Arbeitszeiten zugeordnet werden sollen."
        )
        
        info_label = ctk.CTkLabel(
            container,
            text=info_text,
            justify="left",
            anchor="w",
            wraplength=600
        )
        
        # Drag & Drop Frame
        self.ar_drop_frame = DragDropFrame(
            container,
            callback=self._on_ar_file_selected,
            accepted_files=['.xlsx', '.xls']
        )
        
        # Dateiauswahl-Frame (für Anzeige des ausgewählten Dateinamens)
        file_frame = ctk.CTkFrame(container)
        
        # Label für ausgewählte Datei
        file_label = ctk.CTkLabel(
            file_frame,
            text="Ausgewählte Datei:",
            font=ctk.CTkFont(weight="bold"),
            width=120
        )
        
        # Anzeige des Dateinamens
        self.ar_file_name_label = ctk.CTkLabel(
            file_frame,
            text="Keine Datei ausgewählt",
            anchor="w"
        )
        
        # Mitarbeiter-Auswahl-Frame
        member_frame = ctk.CTkFrame(container)
        
        # Label für Mitarbeiterauswahl
        member_label = ctk.CTkLabel(
            member_frame,
            text="Mitarbeiter:",
            font=ctk.CTkFont(weight="bold"),
            width=120
        )
        
        # Anzeige des ausgewählten Mitarbeiters
        self.selected_member_label = ctk.CTkLabel(
            member_frame,
            text="Kein Mitarbeiter ausgewählt",
            anchor="w"
        )
        
        # Button für Mitarbeiterauswahl
        self.select_member_button = ctk.CTkButton(
            member_frame,
            text="Mitarbeiter auswählen",
            command=self._open_member_selection,
            width=170
        )
        
        # Optionen-Frame
        options_frame = ctk.CTkFrame(container)
        
        # Optionen für den Import
        self.ar_overwrite_var = ctk.BooleanVar(value=True)
        self.ar_overwrite_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Bestehende Einträge überschreiben (gleiche ID und Datum)",
            variable=self.ar_overwrite_var
        )
        
        self.ar_backup_var = ctk.BooleanVar(value=True)
        self.ar_backup_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Backup erstellen",
            variable=self.ar_backup_var
        )
        
        # Import-Button
        self.ar_import_button = ctk.CTkButton(
            container,
            text="Arbeitszeitdaten importieren",
            command=self._import_ar_calendar_data,
            font=ctk.CTkFont(weight="bold"),
            height=40,
            fg_color=("green3", "green4"),
            hover_color=("green4", "green3"),
            state="disabled"  # Initial deaktiviert
        )
        
        # Layout für den Arbeitszeit-Tab
        container.pack(fill="both", expand=True, padx=20, pady=20)
        info_label.pack(anchor="w", pady=(0, 15))
        
        # Drag & Drop Frame
        self.ar_drop_frame.pack(fill="x", pady=(0, 15))
        
        # Dateiauswahl-Frame
        file_frame.pack(fill="x", pady=(0, 15))
        file_label.pack(side="left")
        self.ar_file_name_label.pack(side="left", padx=10, fill="x", expand=True)
        
        # Mitarbeiter-Auswahl-Frame
        member_frame.pack(fill="x", pady=(0, 15))
        member_label.pack(side="left")
        self.selected_member_label.pack(side="left", padx=10, fill="x", expand=True)
        self.select_member_button.pack(side="right")
        
        # Optionen-Frame
        options_frame.pack(fill="x", pady=(0, 20))
        self.ar_overwrite_checkbox.pack(anchor="w", pady=5)
        self.ar_backup_checkbox.pack(anchor="w", pady=5)
        
        # Import-Button
        self.ar_import_button.pack(fill="x", pady=(10, 0))
    
    def _create_projekt_tab(self):
        """Erstellt die Inhalte für den Projektdaten-Tab"""
        # Container-Frame für den Tab
        container = ctk.CTkFrame(self.projekt_tab)
        
        # Info-Text
        info_text = (
            "Importieren Sie Projektdaten aus einer JSON-Datei.\n\n"
            "Die JSON-Datei sollte eine Struktur mit Projekten, Teilnehmern und Sprints enthalten.\n\n"
            "Beispiel:\n"
            '{\n  "projekte": [\n    {\n      "name": "Projektname",\n      "teilnehmer": [...],\n      "sprints": [...]\n    }\n  ]\n}'
        )
        
        info_label = ctk.CTkLabel(
            container,
            text=info_text,
            justify="left",
            anchor="w",
            wraplength=600
        )
        
        # Dateiauswahl-Frame
        file_frame = ctk.CTkFrame(container)
        
        # Label für Dateiauswahl
        file_label = ctk.CTkLabel(
            file_frame,
            text="JSON-Datei:",
            font=ctk.CTkFont(weight="bold"),
            width=80
        )
        
        # Eingabefeld für Dateipfad
        self.projekt_file_entry = ctk.CTkEntry(file_frame, width=400)
        
        # Button für Datei-Browser
        self.projekt_browse_button = ctk.CTkButton(
            file_frame,
            text="Durchsuchen",
            command=lambda: self._browse_file("json", self.projekt_file_entry),
            width=100
        )
        
        # Optionen-Frame
        options_frame = ctk.CTkFrame(container)
        
        # Optionen für den Import
        self.projekt_merge_var = ctk.BooleanVar(value=False)
        self.projekt_merge_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Zusammenführen (bestehende Projekte beibehalten)",
            variable=self.projekt_merge_var
        )
        
        self.projekt_backup_var = ctk.BooleanVar(value=True)
        self.projekt_backup_checkbox = ctk.CTkCheckBox(
            options_frame,
            text="Backup erstellen",
            variable=self.projekt_backup_var
        )
        
        # Import-Button
        self.projekt_import_button = ctk.CTkButton(
            container,
            text="Projektdaten importieren",
            command=self._import_projekt_data,
            font=ctk.CTkFont(weight="bold"),
            height=40,
            fg_color=("green3", "green4"),
            hover_color=("green4", "green3")
        )
        
        # Layout für den Projektdaten-Tab
        container.pack(fill="both", expand=True, padx=20, pady=20)
        info_label.pack(anchor="w", pady=(0, 15))
        
        # Dateiauswahl-Frame
        file_frame.pack(fill="x", pady=(0, 15))
        file_label.pack(side="left")
        self.projekt_file_entry.pack(side="left", padx=10)
        self.projekt_browse_button.pack(side="left")
        
        # Optionen-Frame
        options_frame.pack(fill="x", pady=(0, 20))
        self.projekt_merge_checkbox.pack(anchor="w", pady=5)
        self.projekt_backup_checkbox.pack(anchor="w", pady=5)
        
        # Import-Button
        self.projekt_import_button.pack(fill="x", pady=(10, 0))
    
    def _setup_layout(self):
        """Richtet das Layout der UI-Elemente ein"""
        # Titel-Frame
        self.title_frame.pack(fill="x", padx=20, pady=(20, 10))
        self.title_label.pack(pady=10)
        
        # Status-Frame
        self.status_frame.pack(fill="x", padx=20, pady=(0, 10))
        self.status_label.pack(pady=5, fill="x")
        
        # Tab-View (nimmt den meisten Platz ein)
        self.tab_view.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Button-Frame (unten)
        self.button_frame.pack(fill="x", padx=20, pady=(5, 20))
        self.close_button.pack(side="right", padx=10)
    
    def _on_file_selected(self, filename):
        """
        Wird aufgerufen, wenn eine Datei ausgewählt oder fallen gelassen wurde
        
        Args:
            filename: Pfad zur ausgewählten Datei
        """
        self.selected_file = filename
        self.file_name_label.configure(text=os.path.basename(filename))
        self.bw_tool_import_button.configure(state="normal")
        
        # Status zurücksetzen
        self._show_status("", False)
    
    def _on_ar_file_selected(self, filename):
        """
        Wird aufgerufen, wenn eine Arbeitszeiterfassungsdatei ausgewählt wurde
        
        Args:
            filename: Pfad zur ausgewählten Datei
        """
        self.ar_selected_file = filename
        self.ar_file_name_label.configure(text=os.path.basename(filename))
        
        # Prüfe, ob auch ein Mitarbeiter ausgewählt wurde
        self._update_ar_import_button()
        
        # Status zurücksetzen
        self._show_status("", False)
    
    def _browse_file(self, file_type, entry_widget):
        """
        Öffnet einen Datei-Browser-Dialog
        
        Args:
            file_type: Der Dateityp (csv oder json)
            entry_widget: Das Eingabefeld, in das der Dateipfad eingetragen werden soll
        """
        if file_type.lower() == "csv":
            filetypes = [("CSV-Dateien", "*.csv"), ("Alle Dateien", "*.*")]
            title = "CSV-Datei auswählen"
        elif file_type.lower() == "json":
            filetypes = [("JSON-Dateien", "*.json"), ("Alle Dateien", "*.*")]
            title = "JSON-Datei auswählen"
        else:
            filetypes = [("Alle Dateien", "*.*")]
            title = "Datei auswählen"
        
        # Datei-Dialog öffnen
        filename = filedialog.askopenfilename(
            title=title,
            filetypes=filetypes
        )
        
        # Wenn eine Datei ausgewählt wurde, den Pfad in das Eingabefeld eintragen
        if filename:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, filename)
            
            # Status zurücksetzen
            self._show_status("", False)
    
    def _open_member_selection(self):
        """Öffnet den Dialog zur Auswahl eines Projektmitglieds"""
        try:
            # Lade die Projektdaten
            projects_data = self._load_project_data()
            
            if not projects_data:
                self._show_status("Keine Projektdaten gefunden. Bitte importieren Sie zuerst Projektdaten.", error=True)
                return
            
            # Zeige den Dialog
            dialog = ProjectMemberSelectionDialog(
                self,
                projects_data,
                on_select=self._on_member_selected
            )
            
            # Warte, bis der Dialog geschlossen wird
            self.wait_window(dialog)
        
        except Exception as e:
            self._show_status(f"Fehler beim Öffnen des Mitarbeiterauswahl-Dialogs: {e}", error=True)
    
    def _on_member_selected(self, member_data):
        """
        Wird aufgerufen, wenn ein Projektmitglied ausgewählt wurde
        
        Args:
            member_data: Die Daten des ausgewählten Mitglieds
        """
        self.selected_member = member_data
        
        # Zeige den Namen des ausgewählten Mitglieds an
        member_name = member_data.get("name", "")
        member_id = member_data.get("id", "")
        self.selected_member_label.configure(text=f"{member_name} (ID: {member_id})")
        
        # Prüfe, ob auch eine Datei ausgewählt wurde
        self._update_ar_import_button()
    
    def _update_ar_import_button(self):
        """Aktualisiert den Status des Import-Buttons für die Arbeitszeiterfassung"""
        if self.ar_selected_file and self.selected_member:
            self.ar_import_button.configure(state="normal")
        else:
            self.ar_import_button.configure(state="disabled")
    
    def _import_bw_tool_data(self):
        """Importiert BW Tool Daten aus einer Excel-Datei"""
        # Prüfen, ob eine Datei ausgewählt wurde
        if not self.selected_file:
            self._show_status("Bitte wählen Sie eine Excel-Datei aus.", error=True)
            return
        
        # Prüfen, ob die Datei existiert
        if not os.path.exists(self.selected_file):
            self._show_status(f"Die Datei '{self.selected_file}' existiert nicht.", error=True)
            return
        
        try:
            # Excel-Datei einlesen
            self._show_status("Excel-Datei wird eingelesen...", error=False)
            
            # Ziel-Dateipfad
            target_path = os.path.join("data", "kapa_data.csv")
            
            # Backup erstellen, falls gewünscht
            if self.bw_tool_backup_var.get() and os.path.exists(target_path):
                backup_path = f"{target_path}.bak"
                shutil.copy2(target_path, backup_path)
            
            # Verarbeite die Excel-Datei
            processed_data = self._process_mime_file(self.selected_file)
            
            # Wenn keine Daten verarbeitet wurden
            if not processed_data:
                self._show_status("Keine gültigen Daten in der Datei gefunden.", error=True)
                return
            
            # Einträge überschreiben, falls vorhanden
            if self.bw_tool_overwrite_var.get() and os.path.exists(target_path):
                final_data = self._merge_with_existing_data(processed_data, target_path)
            else:
                # Nur neue Daten hinzufügen
                final_data = processed_data
                if os.path.exists(target_path):
                    with open(target_path, 'r', newline='', encoding='utf-8') as existing_file:
                        reader = csv.reader(existing_file, delimiter=';')
                        existing_data = list(reader)
                        # Nur Zeilen hinzufügen, die noch nicht existieren
                        id_date_pairs = {(row[0], row[1]) for row in existing_data}
                        final_data = [row for row in processed_data if (row[0], row[1]) not in id_date_pairs]
                        final_data = existing_data + final_data
            
            # Daten speichern
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            with open(target_path, 'w', newline='', encoding='utf-8') as output_file:
                writer = csv.writer(output_file, delimiter=';')
                writer.writerows(final_data)
            
            # Erfolgsmeldung anzeigen
            self._show_status(f"Import erfolgreich: {len(processed_data)} Einträge verarbeitet", error=False)
        
        except Exception as e:
            # Fehlermeldung anzeigen
            self._show_status(f"Fehler beim Import: {str(e)}", error=True)
            import traceback
            traceback.print_exc()
    
    def _import_ar_calendar_data(self):
        """Importiert Daten aus einer Arbeitszeiterfassungsdatei"""
        # Prüfen, ob eine Datei und ein Mitarbeiter ausgewählt wurden
        if not self.ar_selected_file:
            self._show_status("Bitte wählen Sie eine Excel-Datei aus.", error=True)
            return
        
        if not self.selected_member:
            self._show_status("Bitte wählen Sie einen Mitarbeiter aus.", error=True)
            return
        
        # Prüfen, ob die Datei existiert
        if not os.path.exists(self.ar_selected_file):
            self._show_status(f"Die Datei '{self.ar_selected_file}' existiert nicht.", error=True)
            return
        
        try:
            # Mitarbeiter-ID extrahieren
            member_id = self.selected_member.get("id", "")
            
            # Excel-Datei verarbeiten
            self._show_status("Arbeitszeitdaten werden verarbeitet...", error=False)
            
            # Ziel-Dateipfad
            target_path = os.path.join("data", "kapa_data.csv")
            
            # Backup erstellen, falls gewünscht
            if self.ar_backup_var.get() and os.path.exists(target_path):
                backup_path = f"{target_path}.bak"
                shutil.copy2(target_path, backup_path)
            
            # Verarbeite die Excel-Datei
            processed_data = self._process_ar_calendar_file(self.ar_selected_file, member_id)
            
            # Wenn keine Daten verarbeitet wurden
            if not processed_data:
                self._show_status("Keine gültigen Daten in der Datei gefunden.", error=True)
                return
            
            # Einträge überschreiben, falls vorhanden
            if self.ar_overwrite_var.get() and os.path.exists(target_path):
                final_data = self._merge_with_existing_data(processed_data, target_path)
            else:
                # Nur neue Daten hinzufügen
                final_data = processed_data
                if os.path.exists(target_path):
                    with open(target_path, 'r', newline='', encoding='utf-8') as existing_file:
                        reader = csv.reader(existing_file, delimiter=';')
                        existing_data = list(reader)
                        # Nur Zeilen hinzufügen, die noch nicht existieren
                        id_date_pairs = {(row[0], row[1]) for row in existing_data}
                        final_data = [row for row in processed_data if (row[0], row[1]) not in id_date_pairs]
                        final_data = existing_data + final_data
            
            # Daten speichern
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            with open(target_path, 'w', newline='', encoding='utf-8') as output_file:
                writer = csv.writer(output_file, delimiter=';')
                writer.writerows(final_data)
            
            # Erfolgsmeldung anzeigen
            self._show_status(f"Import erfolgreich: {len(processed_data)} Einträge verarbeitet", error=False)
        
        except Exception as e:
            # Fehlermeldung anzeigen
            self._show_status(f"Fehler beim Import: {str(e)}", error=True)
            import traceback
            traceback.print_exc()
    
    def _import_projekt_data(self):
        """Importiert Projektdaten aus einer JSON-Datei"""
        # Dateipfad auslesen
        file_path = self.projekt_file_entry.get().strip()
        
        # Prüfen, ob eine Datei ausgewählt wurde
        if not file_path:
            self._show_status("Bitte wählen Sie eine JSON-Datei aus.", error=True)
            return
        
        # Prüfen, ob die Datei existiert
        if not os.path.exists(file_path):
            self._show_status(f"Die Datei '{file_path}' existiert nicht.", error=True)
            return
        
        try:
            # Ziel-Dateipfad
            target_path = os.path.join("data", "project.json")
            
            # Backup erstellen, falls gewünscht
            if self.projekt_backup_var.get() and os.path.exists(target_path):
                backup_path = f"{target_path}.bak"
                shutil.copy2(target_path, backup_path)
            
            # JSON-Datei einlesen
            with open(file_path, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
            
            # Wenn Zusammenführen gewählt wurde, bestehende Daten laden
            if self.projekt_merge_var.get() and os.path.exists(target_path):
                with open(target_path, 'r', encoding='utf-8') as existing_file:
                    existing_data = json.load(existing_file)
                
                # Daten zusammenführen
                if "projekte" in data and "projekte" in existing_data:
                    # Erstelle ein Set mit den Namen der bestehenden Projekte
                    existing_project_names = {p["name"] for p in existing_data["projekte"]}
                    
                    # Füge nur neue Projekte hinzu
                    for project in data["projekte"]:
                        if project["name"] not in existing_project_names:
                            existing_data["projekte"].append(project)
                    
                    # Aktualisierte Daten verwenden
                    data = existing_data
            
            # Daten in die Zieldatei schreiben
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            with open(target_path, 'w', encoding='utf-8') as output_file:
                json.dump(data, output_file, indent=4, ensure_ascii=False)
            
            # Erfolgsmeldung anzeigen
            projekt_count = len(data.get("projekte", []))
            self._show_status(f"Import erfolgreich: {projekt_count} Projekte in '{target_path}'", error=False)
        
        except Exception as e:
            # Fehlermeldung anzeigen
            self._show_status(f"Fehler beim Import: {e}", error=True)
    
    def _show_status(self, message, error=False):
        """
        Zeigt eine Statusmeldung an
        
        Args:
            message: Die anzuzeigende Nachricht
            error: True für eine Fehlermeldung, False für eine Erfolgsmeldung
        """
        if not message:
            self.status_label.configure(text="")
            return
        
        # Nachricht anzeigen
        self.status_label.configure(
            text=message,
            text_color="red" if error else "green"
        )
    
    def _merge_with_existing_data(self, new_data, existing_file_path):
        """
        Führt neue Daten mit vorhandenen Daten zusammen, wobei vorhandene Einträge überschrieben werden
        
        Args:
            new_data: Neue Daten [[ID, DATUM, STUNDEN], ...]
            existing_file_path: Pfad zur bestehenden CSV-Datei
                
        Returns:
            List[List[str]]: Zusammengeführte Daten
        """
        if not os.path.exists(existing_file_path):
            return new_data
        
        try:
            # Vorhandene Daten einlesen
            with open(existing_file_path, 'r', newline='', encoding='utf-8') as file:
                reader = csv.reader(file, delimiter=';')
                existing_data = list(reader)
            
            # Dictionary für schnellere Suche erstellen (ID+DATUM -> Zeilenindex)
            id_date_indices = {(row[0], row[1]): i for i, row in enumerate(existing_data)}
            
            # Neue Daten durchgehen
            for row in new_data:
                key = (row[0], row[1])
                if key in id_date_indices:
                    # Vorhandenen Eintrag überschreiben
                    existing_data[id_date_indices[key]] = row
                else:
                    # Neuen Eintrag hinzufügen
                    existing_data.append(row)
            
            return existing_data
        
        except Exception as e:
            raise Exception(f"Fehler beim Zusammenführen der Daten: {str(e)}")
    
    def _process_mime_file(self, file_path, debug_mode=True):
        """
        Verarbeitet eine MIME-formatierte Datei (z.B. E-Mail oder HTML)
        
        Args:
            file_path: Pfad zur MIME/HTML-Datei
            debug_mode: Aktiviert zusätzliche Debug-Ausgaben
            
        Returns:
            List[List[str]]: Verarbeitete Daten im Format [[ID, DATUM, STUNDEN], ...]
        """
        try:
            # Lese den Dateiinhalt mit verschiedenen Kodierungen
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            except UnicodeDecodeError:
                try:
                    with open(file_path, 'r', encoding='latin1') as f:
                        content = f.read()
                except:
                    with open(file_path, 'r', encoding='cp1252') as f:
                        content = f.read()
            
            # Import regex und datetime
            import re
            import datetime
            
            # Liste für die verarbeiteten Daten
            processed_data = []
            
            # Extrahiere alle Tabellenzeilen
            rows = re.findall(r'<tr.*?>(.*?)</tr>', content, re.DOTALL)
            
            print(f"Gefunden: {len(rows)} Zeilen in der Datei")
            
            # Zähler für diagnostische Zwecke
            cells_too_few = 0
            no_forecast = 0
            processed_rows = 0
            
            for row_idx, row in enumerate(rows):
                # Extrahiere alle Zellen in dieser Zeile
                cells = re.findall(r'<td.*?>(.*?)</td>', row, re.DOTALL)
                
                # Debug-Ausgabe für die ersten 5 Zeilen
                if debug_mode and row_idx < 5:
                    print(f"\nZeile {row_idx+1} hat {len(cells)} Zellen")
                    if len(cells) > 0:
                        for i, cell in enumerate(cells):
                            if i < 15:  # Zeige die ersten 15 Zellen
                                cleaned_cell = re.sub(r'<.*?>', '', cell).replace("&#32;", " ").strip()
                                print(f"  Zelle {i+1}: {cleaned_cell[:30]}")
                
                # Überspringe die Zeile, wenn sie nicht genügend Zellen hat
                if len(cells) < 13:
                    cells_too_few += 1
                    continue
                
                try:
                    # Prüfe, ob die Zeile ein Forecast-Eintrag ist (Spalte 11, Index 10)
                    forecast_cell = cells[10] if len(cells) > 10 else ""
                    forecast_value = re.sub(r'<.*?>', '', forecast_cell).upper().strip()
                    
                    if debug_mode and row_idx < 5:
                        print(f"Zeile {row_idx+1} Forecast-Wert: '{forecast_value}'")
                    
                    if "FCAST" not in forecast_value and "FORECAST" not in forecast_value:
                        no_forecast += 1
                        continue
                    
                    # Name (ID) extrahieren
                    sixth_value = re.sub(r'<.*?>', '', cells[5]).replace("&#32;", " ").strip()
                    if sixth_value == "Fremdleistung OPS" and len(cells) > 6:
                        mitarbeiter_id = re.sub(r'<.*?>', '', cells[6]).replace("&#32;", " ").strip()
                    else:
                        mitarbeiter_id = sixth_value
                    
                    # Datum extrahieren
                    date_cell = cells[7]
                    date_value = re.sub(r'<.*?>', '', date_cell).strip()
                    try:
                        # Versuche, einen Excel-Datumswert zu konvertieren
                        excel_date = int(float(date_value.replace(",", ".")))
                        datetime_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
                        datum = datetime_date.strftime('%d.%m.%Y')
                    except ValueError:
                        # Falls es kein numerischer Wert ist, verwende den Wert direkt
                        datum = date_value
                    
                    # Stunden extrahieren
                    hours_cell = cells[12]
                    hours_value = re.sub(r'<.*?>', '', hours_cell).strip()
                    try:
                        stunden_float = float(hours_value.replace(",", "."))
                        stunden = str(stunden_float)
                    except ValueError:
                        stunden = "0.0"  # Fallback, wenn keine gültige Zahl
                    
                    # Daten zur Liste hinzufügen
                    processed_data.append([mitarbeiter_id, datum, stunden])
                    processed_rows += 1
                    
                    if debug_mode and processed_rows <= 5:
                        print(f"\nErfolgreiche Extraktion aus Zeile {row_idx+1}:")
                        print(f"ID: {mitarbeiter_id}")
                        print(f"Datum: {datum}")
                        print(f"Stunden: {stunden}")
                
                except Exception as e:
                    if debug_mode:
                        print(f"Fehler bei Zeile {row_idx+1}: {e}")
                    continue
            
            print(f"\nZusammenfassung:")
            print(f"Gesamt Zeilen: {len(rows)}")
            print(f"Zeilen mit zu wenigen Zellen: {cells_too_few}")
            print(f"Zeilen ohne Forecast: {no_forecast}")
            print(f"Verarbeitete Zeilen: {processed_rows}")
            
            return processed_data
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"MIME-Verarbeitung fehlgeschlagen: {e}")
            return []
    
    def _process_ar_calendar_file(self, file_path, selected_member_id, debug_mode=True):
        """
        Extrem robuste Verarbeitung einer Excel-Datei im ARCalendarList-Format
        
        Args:
            file_path: Pfad zur Excel-Datei
            selected_member_id: ID des ausgewählten Projektmitglieds
            debug_mode: Aktiviert zusätzliche Debug-Ausgaben
            
        Returns:
            List[List[str]]: Verarbeitete Daten im Format [[ID, DATUM, STUNDEN], ...]
        """
        import pandas as pd
        import re
        import datetime
        import traceback
        import os
        import subprocess
        import tempfile
        import csv
        
        # Liste für die verarbeiteten Daten
        processed_data = []
        temp_files = []
        
        try:
            print(f"Verarbeite ARCalendar-Datei mit robuster Methode: {file_path}")
            print(f"Ausgewählte Mitarbeiter-ID: {selected_member_id}")
            
            # Wir werden mehrere Methoden versuchen
            # Methode 1: Verwende xlwings falls verfügbar
            try:
                print("Versuche Methode 1: xlwings (funktioniert nur auf Windows/Mac mit Excel)...")
                import xlwings as xw
                
                # Erstelle eine temporäre CSV-Datei
                temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
                temp_csv.close()
                temp_files.append(temp_csv.name)
                
                # Starte Excel und konvertiere in CSV
                with xw.App(visible=False) as app:
                    wb = app.books.open(file_path)
                    sheet = wb.sheets[0]
                    
                    # Bereite CSV-Daten vor
                    data = []
                    header = []
                    
                    # Hole die Header
                    for col in range(1, 20):  # Annahme: Nicht mehr als 20 Spalten
                        cell_value = sheet.cells(1, col).value
                        if cell_value is None:
                            break
                        header.append(cell_value)
                    
                    if not header:
                        raise ValueError("Keine Header gefunden")
                    
                    data.append(header)
                    
                    # Hole die Daten
                    for row in range(2, 1000):  # Annahme: Nicht mehr als 1000 Zeilen
                        row_data = []
                        empty_cells = 0
                        
                        for col in range(1, len(header) + 1):
                            cell_value = sheet.cells(row, col).value
                            row_data.append(cell_value)
                            if cell_value is None:
                                empty_cells += 1
                        
                        # Wenn die Zeile fast leer ist, sind wir am Ende
                        if empty_cells >= len(header) - 1:
                            break
                        
                        data.append(row_data)
                    
                    # Schreibe in CSV
                    with open(temp_csv.name, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        for row in data:
                            writer.writerow(row)
                    
                    # Schließe Excel
                    wb.close()
                
                print(f"Datei erfolgreich zu CSV konvertiert: {temp_csv.name}")
                
                # Erstelle ein DataFrame aus der CSV
                df = pd.read_csv(temp_csv.name)
                success = True
            
            except Exception as e:
                print(f"Methode 1 fehlgeschlagen: {str(e)}")
                success = False
            
            # Methode 2: Verwende pyexcel falls verfügbar
            if not success:
                try:
                    print("Versuche Methode 2: pyexcel...")
                    import pyexcel as pe
                    
                    temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
                    temp_csv.close()
                    temp_files.append(temp_csv.name)
                    
                    # Konvertiere von Excel zu CSV
                    pe.save_as(file_name=file_path, dest_file_name=temp_csv.name)
                    
                    print(f"Datei erfolgreich zu CSV konvertiert: {temp_csv.name}")
                    
                    # Erstelle ein DataFrame aus der CSV
                    df = pd.read_csv(temp_csv.name)
                    success = True
                
                except Exception as e:
                    print(f"Methode 2 fehlgeschlagen: {str(e)}")
                    success = False
            
            # Methode 3: Verwende LibreOffice/OpenOffice falls installiert
            if not success:
                try:
                    print("Versuche Methode 3: LibreOffice/OpenOffice...")
                    
                    temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
                    temp_csv.close()
                    temp_files.append(temp_csv.name)
                    
                    # Prüfe auf LibreOffice/OpenOffice
                    libreoffice_paths = [
                        '/Applications/LibreOffice.app/Contents/MacOS/soffice',  # macOS
                        'C:\\Program Files\\LibreOffice\\program\\soffice.exe',  # Windows
                        'C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe',  # Windows 32-bit
                        '/usr/bin/libreoffice',  # Linux
                        '/usr/bin/soffice'  # Linux
                    ]
                    
                    soffice_path = None
                    for path in libreoffice_paths:
                        if os.path.exists(path):
                            soffice_path = path
                            break
                    
                    if soffice_path:
                        # Konvertiere mit LibreOffice/OpenOffice
                        cmd = [
                            soffice_path,
                            '--headless',
                            '--convert-to', 'csv',
                            '--outdir', os.path.dirname(temp_csv.name),
                            file_path
                        ]
                        
                        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        stdout, stderr = process.communicate()
                        
                        if process.returncode != 0:
                            raise Exception(f"LibreOffice Konvertierung fehlgeschlagen: {stderr.decode('utf-8')}")
                        
                        # Finde die generierte CSV-Datei
                        base_name = os.path.basename(file_path)
                        csv_name = os.path.splitext(base_name)[0] + '.csv'
                        generated_csv = os.path.join(os.path.dirname(temp_csv.name), csv_name)
                        
                        # Kopiere die generierte Datei an den erwarteten Ort
                        import shutil
                        shutil.copy2(generated_csv, temp_csv.name)
                        
                        # Lösche die generierte Datei
                        os.remove(generated_csv)
                        
                        print(f"Datei erfolgreich zu CSV konvertiert: {temp_csv.name}")
                        
                        # Erstelle ein DataFrame aus der CSV
                        df = pd.read_csv(temp_csv.name)
                        success = True
                    else:
                        raise Exception("LibreOffice/OpenOffice nicht gefunden")
                
                except Exception as e:
                    print(f"Methode 3 fehlgeschlagen: {str(e)}")
                    success = False
            
            # Methode 4: Manueller Fallback - mit vereinfachtem openpyxl
            if not success:
                try:
                    print("Versuche Methode 4: Vereinfachtes openpyxl...")
                    
                    temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
                    temp_csv.close()
                    temp_files.append(temp_csv.name)
                    
                    from openpyxl import load_workbook
                    from openpyxl.utils.exceptions import InvalidFileException
                    
                    try:
                        # Versuche zu laden, ignoriere Stile
                        wb = load_workbook(filename=file_path, read_only=True, data_only=True, keep_links=False)
                    except TypeError:
                        # Fallback für ältere openpyxl-Versionen
                        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                    
                    ws = wb.active
                    
                    # Schreibe in CSV
                    with open(temp_csv.name, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        for row in ws.iter_rows(values_only=True):
                            writer.writerow(row)
                    
                    print(f"Datei erfolgreich zu CSV konvertiert: {temp_csv.name}")
                    
                    # Erstelle ein DataFrame aus der CSV
                    df = pd.read_csv(temp_csv.name)
                    success = True
                
                except Exception as e:
                    print(f"Methode 4 fehlgeschlagen: {str(e)}")
                    success = False
            
            # Methode 5: Manuelle Extraktion mit geringster Abhängigkeit
            if not success:
                try:
                    print("Versuche Methode 5: Extrem robuste manuelle Extraktion...")
                    
                    # Erstelle eine Dummy-CSV mit Platzhalterwerten
                    temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
                    temp_files.append(temp_csv.name)
                    
                    with open(temp_csv.name, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(["Date", "TotalHours"])  # Kopfzeile
                        
                        # Rohbytes aus der Excel-Datei extrahieren
                        with open(file_path, 'rb') as excel_file:
                            excel_data = excel_file.read()
                        
                        # Suche nach Datums- und Zahlenwerten in den Rohbytes
                        date_patterns = [
                            rb'\d{2}[/.-]\d{2}[/.-]\d{4}',  # DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY
                            rb'\d{4}[/.-]\d{2}[/.-]\d{2}'   # YYYY/MM/DD, YYYY-MM-DD, YYYY.MM.DD
                        ]
                        
                        # Finde alle möglichen Datumsangaben
                        all_dates = []
                        for pattern in date_patterns:
                            matches = re.findall(pattern, excel_data)
                            all_dates.extend([m.decode('utf-8', errors='ignore') for m in matches])
                        
                        # Finde alle möglichen Zahlen (für Stunden)
                        all_numbers = re.findall(rb'\b\d+\.\d+\b|\b\d+,\d+\b|\b\d+\b', excel_data)
                        all_numbers = [n.decode('utf-8', errors='ignore') for n in all_numbers]
                        
                        # Wenn wir genügend Daten haben, schreiben wir sie in die CSV
                        if all_dates and all_numbers:
                            # Filtere die Zahlen, um nur solche zwischen 0.1 und 24 zu behalten (plausible Arbeitsstunden)
                            filtered_numbers = []
                            for num in all_numbers:
                                try:
                                    value = float(num.replace(',', '.'))
                                    if 0.1 <= value <= 24:
                                        filtered_numbers.append(num)
                                except ValueError:
                                    continue
                            
                            # Nehme die kleinere Anzahl von Daten
                            rows_count = min(len(all_dates), len(filtered_numbers))
                            
                            # Schreibe Zeilen
                            for i in range(rows_count):
                                writer.writerow([all_dates[i], filtered_numbers[i]])
                        else:
                            raise ValueError("Keine Datums- oder Zahlenwerte gefunden")
                    
                    print(f"Datei erfolgreich zu CSV extrahiert: {temp_csv.name}")
                    
                    # Erstelle ein DataFrame aus der CSV
                    df = pd.read_csv(temp_csv.name)
                    success = True
                
                except Exception as e:
                    print(f"Methode 5 fehlgeschlagen: {str(e)}")
                    success = False
            
            # Wenn keine Methode funktioniert hat
            if not success:
                error_message = "Alle Methoden zur Excel-Verarbeitung sind fehlgeschlagen."
                print(error_message)
                self._show_status(error_message, error=True)
                return []
            
            print(f"Excel-Datei erfolgreich gelesen: {df.shape[0]} Zeilen, {df.shape[1]} Spalten")
            
            # Zeige die Spalten zur Überprüfung
            if debug_mode:
                print("\nSpaltennamen:")
                print(df.columns.tolist())
                print("\nErste 5 Zeilen der CSV-Datei:")
                print(df.head().to_string())
            
            # Überprüfe, ob die erwarteten Spalten vorhanden sind
            required_columns = ["Date", "TotalHours"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            # Versuche alternative Spaltennamen, falls die Standardnamen nicht gefunden wurden
            alternative_columns = {}
            if missing_columns:
                print(f"Warnung: Folgende Spalten fehlen: {missing_columns}")
                print("Suche nach alternativen Spaltennamen...")
                
                # Suche nach Spalten mit ähnlichen Namen
                column_mapping = {
                    "Date": ["date", "datum", "tag", "day", "datum", "arbeitszeit", "zeitraum", "day", "date1"],
                    "TotalHours": ["total", "hours", "stunden", "gesamtstunden", "zeit", "summe", "gesamt", "hours", "totalhours", "total"]
                }
                
                for required_col, alternatives in column_mapping.items():
                    if required_col in missing_columns:
                        for alt in alternatives:
                            matching_cols = [col for col in df.columns if isinstance(col, str) and alt.lower() in col.lower()]
                            if matching_cols:
                                alternative_columns[required_col] = matching_cols[0]
                                print(f"Alternative für '{required_col}' gefunden: '{matching_cols[0]}'")
                                break
            
            # Aktualisiere die Spaltennamen mit den gefundenen Alternativen
            date_column = alternative_columns.get("Date", "Date")
            hours_column = alternative_columns.get("TotalHours", "TotalHours")
            
            # Wenn immer noch keine Spalten gefunden wurden, versuche numerische Indizes
            if date_column not in df.columns or hours_column not in df.columns:
                print("Keine passenden Spaltennamen gefunden, versuche mit Position der Spalten")
                # Zeige alle verfügbaren Spalten
                for i, col in enumerate(df.columns):
                    print(f"Spalte {i}: {col}")
                
                # Versuche die ersten beiden Spalten zu verwenden, wenn sie nicht gefunden wurden
                if date_column not in df.columns and len(df.columns) > 0:
                    date_column = df.columns[0]
                    print(f"Verwende erste Spalte als Datumsspalte: {date_column}")
                
                if hours_column not in df.columns and len(df.columns) > 1:
                    hours_column = df.columns[1]
                    print(f"Verwende zweite Spalte als Stundenspalte: {hours_column}")
            
            # Überprüfe erneut, ob die Spalten jetzt verfügbar sind
            if date_column not in df.columns:
                print(f"Fehler: Keine geeignete Spalte für Datum gefunden!")
                self._show_status("Konnte keine Datumsspalte in der CSV-Datei identifizieren.", error=True)
                return []
            
            if hours_column not in df.columns:
                print(f"Fehler: Keine geeignete Spalte für Stunden gefunden!")
                self._show_status("Konnte keine Stundenspalte in der CSV-Datei identifizieren.", error=True)
                return []
            
            # Verarbeite die Daten
            processed_rows = 0
            
            for index, row in df.iterrows():
                try:
                    # Extrahiere das Datum
                    date_value = row[date_column]
                    
                    # Konvertiere das Datum ins richtige Format
                    if isinstance(date_value, (pd.Timestamp, datetime.datetime, datetime.date)):
                        # Pandas Timestamp direkt formatieren
                        datum = date_value.strftime('%d.%m.%Y')
                    else:
                        try:
                            # Versuche, den Wert als Datum zu parsen
                            datum_obj = pd.to_datetime(date_value)
                            datum = datum_obj.strftime('%d.%m.%Y')
                        except:
                            # Wenn das Parsen fehlschlägt, versuche zu prüfen, ob es ein Datums-String ist
                            date_str = str(date_value)
                            if re.match(r'\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}', date_str) or re.match(r'\d{4}[/.-]\d{1,2}[/.-]\d{1,2}', date_str):
                                # Versuche das Format zu bestimmen und zu konvertieren
                                formats = ['%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']
                                converted = False
                                for fmt in formats:
                                    try:
                                        datum_obj = datetime.datetime.strptime(date_str, fmt)
                                        datum = datum_obj.strftime('%d.%m.%Y')
                                        converted = True
                                        break
                                    except:
                                        continue
                                
                                if not converted:
                                    datum = date_str
                            else:
                                datum = date_str
                    
                    # Extrahiere die Stunden
                    hours_value = row[hours_column]
                    
                    # Konvertiere die Stunden in eine Zahl
                    if pd.isna(hours_value):
                        # Überspringe Zeilen ohne Stundenwert
                        continue
                    
                    try:
                        # Konvertiere in float und dann in String
                        if isinstance(hours_value, str):
                            # Kommas durch Punkte ersetzen
                            hours_value = hours_value.replace(',', '.')
                        
                        stunden = str(float(hours_value))
                    except ValueError:
                        # Wenn die Konvertierung fehlschlägt, überspringe die Zeile
                        print(f"Ungültiger Stundenwert in Zeile {index+2}: {hours_value}")
                        continue
                    
                    # Füge die Daten zur Liste hinzu
                    processed_data.append([selected_member_id, datum, stunden])
                    processed_rows += 1
                    
                    # Debug-Ausgabe für erfolgreiche Extraktion
                    if debug_mode and processed_rows <= 5:
                        print(f"\nErfolgreiche Extraktion aus Zeile {index+2}:")
                        print(f"ID: {selected_member_id}")
                        print(f"Datum: {datum}")
                        print(f"Stunden: {stunden}")
                
                except Exception as e:
                    print(f"Fehler beim Verarbeiten von Zeile {index+2}: {e}")
                    if debug_mode:
                        traceback.print_exc()
                    continue
            
            print(f"\nVerarbeitungsstatistik:")
            print(f"Gesamt Zeilen in CSV: {len(df)}")
            print(f"Erfolgreich verarbeitete Zeilen: {processed_rows}")
            print(f"Extrahierte Datensätze: {len(processed_data)}")
            
            return processed_data
        
        except Exception as e:
            error_message = f"Allgemeiner Fehler bei der Verarbeitung: {e}"
            print(error_message)
            traceback.print_exc()
            self._show_status(error_message, error=True)
            return []
        
        finally:
            # Aufräumen: Temporäre Dateien löschen
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    try:
                        print(f"Lösche temporäre Datei: {temp_file}")
                        os.remove(temp_file)
                    except Exception as e:
                        print(f"Warnung: Konnte temporäre Datei nicht löschen: {e}")
    
    def _load_project_data(self):
        """
        Lädt die Projektdaten aus der project.json-Datei
        
        Returns:
            dict: Die geladenen Projektdaten oder None bei einem Fehler
        """
        try:
            # Pfad zur project.json
            json_path = os.path.join("data", "project.json")
            
            # Prüfe, ob die Datei existiert
            if not os.path.exists(json_path):
                print(f"Warnung: Datei {json_path} nicht gefunden.")
                return None
            
            # Lade die Daten
            with open(json_path, 'r', encoding='utf-8') as json_file:
                return json.load(json_file)
        
        except Exception as e:
            print(f"Fehler beim Laden der Projektdaten: {e}")
            return None
    
    def center_window(self):
        """Zentriert das Fenster auf dem Bildschirm"""
        self.update_idletasks()
        
        # Fenstergröße
        width = self.winfo_width()
        height = self.winfo_height()
        
        # Bildschirmgröße
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Position berechnen
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # Fenster positionieren
        self.geometry(f"{width}x{height}+{x}+{y}")